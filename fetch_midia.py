# -*- coding: utf-8 -*-
"""
Puxa os dados de campanha da planilha que o Jose exporta do Meta pelo Stract
e grava midia.json.

Fonte unica de verdade para numeros de campanha: a planilha, nao o PostHog.
Corte em 19/08/2026 - foi quando o gatilho de checkout mudou. Antes disso
os numeros nao sao comparaveis.
"""
import io
import json
import os
import sys
import unicodedata
import urllib.request
from collections import defaultdict
from datetime import datetime, timezone

import openpyxl

PLANILHA = os.environ.get(
    "PLANILHA_ID", "1VddMyaSUayN9dhWB7yVgmWk0efDc9G-mQCFUxLv3pJU")
CORTE  = os.environ.get("CORTE_MIDIA", "2026-08-19")
PREFIXO = os.environ.get("PREFIXO_CAMPANHA", "EN_")
MOEDA  = os.environ.get("MOEDA", "USD")
URL = f"https://docs.google.com/spreadsheets/d/{PLANILHA}/export?format=xlsx"


def norm(s):
    s = unicodedata.normalize("NFKD", str(s or ""))
    s = "".join(c for c in s if not unicodedata.combining(c))
    return " ".join(s.lower().split())


def aba(wb, alvo):
    for nome in wb.sheetnames:
        if norm(nome) == norm(alvo):
            return wb[nome]
    raise SystemExit(f"aba nao encontrada: {alvo} (tem: {wb.sheetnames})")


def num(v):
    try:
        return float(v)
    except (TypeError, ValueError):
        return 0.0


def ler(ws):
    """Devolve as linhas da aba como dicts, ja filtradas por campanha e data."""
    linhas = ws.iter_rows(values_only=True)
    hdr = [norm(c) for c in next(linhas)]
    out = []
    for r in linhas:
        if not r or not r[0]:
            continue
        d = dict(zip(hdr, r))
        data = str(r[0])[:10]
        camp = str(d.get("campanha") or "")
        if not camp.startswith(PREFIXO) or data < CORTE:
            continue
        d["_d"] = data
        out.append(d)
    return out


def agrega(linhas, chave=None):
    """Soma as metricas por (data[, chave])."""
    campos = [("inv", "investimento"), ("imp", "impressoes"), ("clk", "cliques"),
              ("vis", "visitas no site"), ("ckt", "checkout"),
              ("vendas", "vendas"), ("fat", "faturamento")]
    acc = defaultdict(lambda: defaultdict(float))
    for r in linhas:
        k = (r["_d"],) if chave is None else (r["_d"], str(r.get(chave) or "-"))
        for curto, col in campos:
            acc[k][curto] += num(r.get(col))
    saida = []
    for k in sorted(acc):
        item = {"d": k[0]}
        if chave is not None:
            item[chave[:3]] = k[1]
        item.update({c: round(v, 2) for c, v in acc[k].items()})
        saida.append(item)
    return saida


def main():
    print(f"baixando planilha {PLANILHA} ...", file=sys.stderr)
    with urllib.request.urlopen(URL, timeout=180) as r:
        buf = io.BytesIO(r.read())
    wb = openpyxl.load_workbook(buf, read_only=True, data_only=True)

    dia_raw = ler(aba(wb, "Meta Ads"))
    pub_raw = ler(aba(wb, "Meta Ads - Publico"))

    if not dia_raw:
        raise SystemExit(
            f"nenhuma linha com campanha {PREFIXO}* a partir de {CORTE}. "
            "A planilha foi atualizada?")

    dados = {
        "corte": CORTE,
        "moeda": MOEDA,
        "campanhas": sorted({str(r.get("campanha")) for r in dia_raw}),
        "dia": agrega(dia_raw),
        "criativo": agrega(dia_raw, "criativo"),
        "publico": agrega(pub_raw, "publico"),
        "atualizado_em": datetime.now(timezone.utc).isoformat(timespec="seconds"),
    }

    with open("midia.json", "w", encoding="utf-8") as f:
        json.dump(dados, f, ensure_ascii=False, separators=(",", ":"))

    t = {c: sum(x.get(c, 0) for x in dados["dia"]) for c in
         ("inv", "imp", "clk", "vis", "ckt", "vendas")}
    print(f"midia.json: {len(dados['dia'])} dias, "
          f"{len(dados['criativo'])} linhas de criativo, "
          f"{len(dados['publico'])} de publico", file=sys.stderr)
    print(f"  {CORTE} em diante | invest {t['inv']:.2f} {MOEDA} | "
          f"{t['imp']:.0f} impressoes | {t['clk']:.0f} cliques | "
          f"{t['vis']:.0f} visitas | {t['ckt']:.0f} checkouts | "
          f"{t['vendas']:.0f} vendas", file=sys.stderr)


if __name__ == "__main__":
    main()
